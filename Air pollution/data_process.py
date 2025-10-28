# -*- coding: utf-8 -*-
"""
Created on Thu Jun  5 13:09:36 2025

@author: shail
"""


import openpyxl
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import pyexcel
import datetime
import dateutil.parser as parser
import matplotlib.pyplot as plt
import seaborn as sns


path = 'E:\\Air\\Air pollution\\Annual'
months_dict = {'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6, 'july': 7, 'august': 8, 'september': 9,
               'october': 10, 'november': 11, 'december': 12}



def get_air_file_month(directory):

    files = [f for f in os.listdir(directory) if os.path.isfile(
        os.path.join(directory, f))]
    #print(files)
    for file in files:
        if 'air' in file.lower():
            #print ("air found")
            return directory + '\\' + file


def get_all_air_files_for_year(year):

    year_str = str(year)
    year_path = path+'\\' + year_str
    directory = year_path
    subfolders = [os.path.join(directory, f) for f in os.listdir(
        directory) if os.path.isdir(os.path.join(directory, f))]
    print(subfolders)

    month_file_map = {}
    for sub_folder in subfolders:
       # print("current_subfolder=",sub_folder )
        month_year = sub_folder.split("\\")[-1]
        if year_str in month_year:
            month = month_year.split(year_str)[0].strip()
            print(month)
        else:
            if year_str[2:] in month_year:
                month = month_year.split(year_str[2:])[0].strip()
                print("found 2 digit year", month_year)

            else:
                continue

        #month = month_year.split(' ')[0]
        if month.lower() not in months_dict:
            continue
        #print( month, months_dict[month.lower()])
        file_for_month = get_air_file_month(sub_folder)
        month_file_map[months_dict[month.lower()]] = file_for_month
    return month_file_map


def find_air_file_for_month(subfolders, month):
    #months = month_dict.keys()
    # for month in months:

    for folder in subfolders:
        if month in folder.lower():
            # month_folder_list.append(folder)
            air_file = get_air_file_month(folder)
            if air_file == None:
                print(folder, "has no air file")
                continue
            else:
                print("found air file for ", month)
                return air_file


def map_directory_to_months_for_year(year):
    year_str = str(year)
    year_path = path+'\\' + year_str
    directory = year_path
    subfolders = [os.path.join(directory, f) for f in os.listdir(
        directory) if os.path.isdir(os.path.join(directory, f))]
    # print(subfolders)
    month_file_map = {}
    for month in months_dict.keys():
        air_file = find_air_file_for_month(subfolders, month)
        if air_file is not None:
            month_num = months_dict[month]
            month_file_map[month] = air_file
    return month_file_map


#output = map_directory_to_months_for_year(2017)


num_so2_cols = 5
num_nox_cols = 6


#freq = get_mode_freq(my_data)
def isNotDate_old(cell_val):
    if isinstance(cell_val, datetime.datetime):
        return False
    cell_val = str(cell_val)
    slash_list = cell_val.split("/")
    if(len(slash_list) < 3):
        # " corrected=", date_time )
        print("!!!!incorrent date format", cell_val)
        return True
    if(len(slash_list) == 3):
        date_time = parser.parse(cell_val)

        print("original", cell_val, "converted", date_time)
    return False


def isNotDate(cell_val, year, month):
    if isinstance(cell_val, datetime.datetime):
        return False
    cell_val = str(cell_val)
    separator = None
    if '.' in cell_val:
        separator = '.'
    if '/' in cell_val:
        separator = '/'
    slash_list = cell_val.split(separator)
    if(len(slash_list) != 3):
        print("!!!!incorrent date format", cell_val, "in",
              year, month)  # " corrected=", date_time )
        return True
    # if(len(slash_list)==3):
    #    date_time = parser.parse(cell_val)
    #    if(separator=='.'):
    #        print("original", cell_val, "converted", date_time )
    return False


def get_date_yyyy_mm_dd(slash_list, year, month):
    if(len(slash_list) != 3):
        print("date does not have 3 numbers", slash_list)
        return [None, None, None]
    two_digits_of_year = year-2000
    first_num = int(slash_list[0])
    extracted_month = int(slash_list[1])
    third_num = int(slash_list[2])
    if(first_num < 100) and (third_num < 100):
        if (first_num == third_num):
            return[year, extracted_month, third_num]
        if (first_num == two_digits_of_year) and (third_num != two_digits_of_year):
            return[year, extracted_month, third_num]

        if (first_num != two_digits_of_year) and (third_num == two_digits_of_year):
            return[year, extracted_month, first_num]
    if first_num > 2000:
        return[first_num, extracted_month, third_num]
    if third_num > 2000:
        return[third_num, extracted_month, first_num]
    print("date  has issue", slash_list)
    return[None, None, None]


def extract_correct_date_new(date_str, year, month):
    running_month = int(months_dict[month])
    extracted_day = None
    extracted_month = None
    extracted_year = None
    cell_val = str(date_str)
    if isinstance(date_str, datetime.datetime):
        extracted_day = date_str.day
        extracted_month = date_str.month
        extracted_year = date_str.year
    else:
        #print('new date=', cell_val)
        separator = None
        if '.' in cell_val:
            separator = '.'
        if '/' in cell_val:
            separator = '/'
        slash_list = cell_val.split(separator)
        [extracted_year, extracted_month, extracted_day] = get_date_yyyy_mm_dd(
            slash_list, year, month)
    if(extracted_year == None):
        return None
    if (extracted_day != running_month) and (extracted_month != running_month):
        print("#### correct month not present in ",
              cell_val, "for", year, month)
        return None
    else:
        if(extracted_day == running_month):
            extracted_day = extracted_month
    correct_date = None
    if(extracted_year < 100) and (separator == '.'):
        #print("******two digit year", extracted_year, "in", cell_val, "for" ,year, month)
        extracted_year = 2000+year

    try:
        datetime.datetime(extracted_year, running_month, extracted_day)
    except ValueError:
        print("$$$$$$ not valid date", date_str, "in", year, month)
    else:
        correct_date = datetime.datetime(
            extracted_year, running_month, extracted_day)

    return correct_date


def extract_correct_date(date_str, year, month):
    cell_val = str(date_str)
    #print('new date=', cell_val)
    separator = None
    if '.' in cell_val:
        separator = '.'
    if '/' in cell_val:
        separator = '/'
    slash_list = cell_val.split(separator)
    correct_date = None
    if isinstance(date_str, datetime.datetime):
        separator = '-'
        dash_list = cell_val.split(' ')[0].split('-')
        extracted_year = int(dash_list[0])

        extracted_month = int(dash_list[1])
        extracted_day = int(dash_list[2])
        #given_date = parser.parse(date_str)
        #extracted_month = date_str.month
        if(extracted_month != int(months_dict[month])):
            if(extracted_day == int(months_dict[month])):
                # format changed from YYYY-MM_DD tp YYYY-DD-MM

                extracted_day = extracted_month

            else:
                #correct_date= date_str.replace(month=int(months_dict[month]))
                print("changing month in", date_str, "to",
                      month, correct_date, " in year", year)
                try:
                    datetime.datetime(
                        extracted_year, months_dict[month], extracted_day)
                except ValueError:
                    print("$$$$$$ not valid date", date_str, "in", year, month)
                else:
                    correct_date = datetime.datetime(
                        extracted_year, months_dict[month], extracted_day)

            return correct_date
        else:
            #datetime.datetime(extracted_year,extracted_month, extracted_day)
            return date_str

        # return date_str
    # if (months_dict[month])!= int(slash_list[1]):
    #    print("!!! incorrect month in", date_str,  " detected month=",slash_list[1], "in", year, month)
    #correct_date = datetime.datetime(year,int(slash_list[1]), int(slash_list[0]))
    try:
        datetime.datetime(year, int(months_dict[month]), int(slash_list[0]))
    except ValueError:
        print("$$$$$$ not valid date", date_str, "in", year, month)
    else:
        correct_date = datetime.datetime(
            year, int(months_dict[month]), int(slash_list[0]))
    return correct_date


def fill_cells(my_data, year, month):
    index_list = my_data.index
    date_col = my_data[my_data.columns[0]].copy()
    date_col = date_col.dropna()
    gap_list = []
    #print("col_size after dropping na",date_col.size)
    for a in range(1, date_col.size):
        gap = date_col.index[a] - date_col.index[a-1]
        gap_list.append(gap)
    gap_series = pd.Series(gap_list)
    val_counts = gap_series.value_counts()
    sorted_val_counts = val_counts.sort_values(ascending=False)
    #periodicity = sorted_val_counts[0]
    #mode = gap_series.mode()
    # print("periods=",sorted_val_counts)
    prediocity = sorted_val_counts.index[0]
    new_date = None
    i = 0
    for index in index_list:
        current_date = my_data[my_data.columns[0]][index]
        if(current_date is None) or (str(current_date) == 'nan') or isNotDate(current_date, year, month):
            if(i < prediocity):
                #print("i=", i, " found empty cell")
                my_data.loc[index, my_data.columns[0]] = new_date
                #print("new val=",my_data.loc[index,my_data.columns[0]])

        else:
            #print( "found new date:", current_date, "i=", i)
            i = 0
            #new_date= parser.parse(str(current_date))
            new_date = extract_correct_date_new(current_date, year, month)
            my_data.loc[index, my_data.columns[0]] = new_date

        i += 1
    return my_data

    # periodicity=get_mode_freq(data)


def create_location_compound_df(my_excel, location, start_index, compund, year, month):
    numcols = 0
    col_names = []
    if compund == 'so2':
        numcols = num_so2_cols
        #print("in so2", numcols)
        col_names = ['Date', 'Abs.',
                     'V(m3)', 'Conc.', 'Aveg.', 'compound', 'location']

    if compund == 'nox':
        numcols = num_nox_cols
        col_names = ['Date', 'Abs.', 'Conc.', 'V(m3)', 'NOx conc', 'Average', 'compound',
                     'location']
    last_index = start_index + numcols
#    print("in",compund )
    #print ("start", start_index, " last", last_index)
    data = my_excel.iloc[2:, start_index: last_index]
    #print ( data.shape)
    data.columns = data.iloc[0]
    data.drop(index=2, axis=0, inplace=True)
    cols = data.columns
    for col in cols[1:numcols]:
        data[col] = pd.to_numeric(data[col], errors="coerce")
    data.dropna(axis=0, how='all', inplace=True)

    data['compound'] = compund
    data['location'] = location
    data.reset_index(inplace=True)
    data = data.drop(columns=['index'])
    data = fill_cells(data, year, month)
    data.columns = col_names
    data['year'] = year
    data['month'] = month
    return data


# os.chdir(path)
#year ='2024'
#month = 'january'

# Load the workbook and select the sheet
#workbook = load_workbook("D:\\Air Pollution Sonawavne Sir\\SAMP Data\\Annual\\2024\\jANUARY 2024\\Template_Air.xlsx")
#sheet = workbook.active
#my_excel = pd.DataFrame(workbook[workbook.sheetnames[0]].values)


def process_air_excel(file_path, year, month):
    if not file_path or not os.path.exists(file_path):
        print(" No valid Excel file for {year} {month} → {file_path}")
        return {"so2": [], "nox": []}   # return empty instead of crashing

    if file_path.endswith(".xls"):
        xls = pd.ExcelFile(file_path, engine="xlrd")
    elif file_path.endswith(".xlsx"):
        xls = pd.ExcelFile(file_path, engine="openpyxl")
    else:
        print("Unsupported file type: {file_path}")
        return {"so2": [], "nox": []}
    sheets = xls.sheet_names

    my_excel = pd.read_excel(xls, sheets[0], header=None)

    my_excel.dropna(axis=0, how='all', inplace=True)
    #my_excel.dropna(axis=1, how='all', inplace=True)
    row0_data = my_excel.iloc[0].dropna()
    so2_start_column = row0_data.index[0]
    nox_start_column = row0_data.index[1]
    loc_data = my_excel.iloc[1].dropna()

    so2_list = []
    nox_list = []
    for index in loc_data.index:
        location = loc_data[index]
        location = location.strip()
        if index < nox_start_column:
            compound = 'so2'
            #print ("processing so2 for " ,location)
            data = create_location_compound_df(
                my_excel, location, index, 'so2', year, month)
            so2_list.append(data)
        else:
            compound = 'nox'
            #print ("processing nox for " ,location)

            data = create_location_compound_df(
                my_excel, location, index, 'nox', year, month)
            nox_list.append(data)
        #print(index,location, compound)
    outmap = {}
    outmap['so2'] = so2_list
    outmap['nox'] = nox_list
    return outmap


# get all air file paths
year_file_list_map = {}
for year in range(2010, 2026):
    month_file_map = map_directory_to_months_for_year(year)
    year_file_list_map[year] = month_file_map

#print (month_file_map)

# process all files

result_list = []
for year in year_file_list_map.keys():
    month_file_map = year_file_list_map[year]
    for month in month_file_map.keys():
        #print("processing air file for", year, month )
        filepath = month_file_map[month]
        outmap = process_air_excel(filepath, year, month)
        result_list.append(outmap)

i = 0
for my_map in result_list:
    so2_list = my_map['so2']
    nox_list = my_map['nox']

    if not so2_list or not nox_list:
        continue   # skip empty months to avoid error

    if i == 0:
        so2_df = pd.concat(so2_list, ignore_index=True)
        nox_df = pd.concat(nox_list, ignore_index=True)
        i += 1
    else:
        tmp_so2 = pd.concat(so2_list, ignore_index=True)
        tmp_nox = pd.concat(nox_list, ignore_index=True)
        so2_df = pd.concat([so2_df, tmp_so2], ignore_index=True)
        nox_df = pd.concat([nox_df, tmp_nox], ignore_index=True)
        i += 1



def check_year_Correctness(row):

    if(not isinstance(row['Date'], datetime.datetime)):
        print(" date ", row['Date'], row['year'], row['month'])
        return None
    # date=row['Date'].year

    return row['year'] == row['Date'].year


so2_df1 = so2_df.dropna(subset=['Date'])

so2_df1['year_match'] = so2_df1.apply(check_year_Correctness, axis=1)


so2_df2 = so2_df1[so2_df1['year_match'] == True]

so2_df2_2012_plus = so2_df2[so2_df2['year'] > 2010]

nox_df2 = nox_df.dropna(subset=['Date'])
nox_df2['year_match'] = nox_df2.apply(check_year_Correctness,axis = 1)
nox_df3 = nox_df2[nox_df2['year_match']==True]
nox_df2_2011_plus = nox_df3[nox_df3['year']>2010]
nox_df2_2011_plus.to_excel("Nox_2011_plus.xlsx",index=False)
print("******************************************NOx df doneee****************************************************")



print("-----------------------------------------------------------------------------------------------")
def extract_year(row):
    return row['Date'].year


def extract_month(row):
    return row['Date'].month


def get_iqr_outliers(df, col_name, threshold=3):
    df = df.dropna()
    data = df[col_name]
    q1 = data.quantile(0.25)
    q3 = data.quantile(0.75)
    iqr = q3-q1
    lower_bound = q1-threshold*iqr
    upper_bound = q3+threshold*iqr
    condition = (data < lower_bound) | (data > upper_bound)
    outliers = df[condition]
    return outliers


def mark_outliers(df, col_name, threshold=3):
    df['is_outlier'] = False
    outliers = get_iqr_outliers(df, col_name, threshold)
    for index in outliers.index:
        df.loc[index, 'is_outlier'] = True
    return df


def location_wise_data(mydata):
    location_data_map = {}
    locations = mydata.location.unique()
    for location in locations:
        location_data_map[location] = mydata[mydata.location == location]
    return location_data_map


def create_date_for_month(row):
    year = row['year']
    month1 = row['month']
    #month = months_dict[month1]

    return datetime.datetime(year, month1, 1)


def aggregate_for_month(mydata):
    monthly_data = mydata.groupby(['year', 'month', 'location'])[
        'mean'].agg([np.sum, np.mean, np.std, np.median])
    monthly_data = monthly_data.reset_index()
    monthly_data['month_date'] = monthly_data.apply(
        create_date_for_month, axis=1)
    return monthly_data


def create_day_level(mydata):

    day_loc_conc_avg = mydata.groupby(['Date', 'location'])[
        'Conc.'].agg([np.sum, np.mean, np.std])
    day_loc_conc_avg1 = day_loc_conc_avg.reset_index()
    day_loc_conc_avg1['year'] = day_loc_conc_avg1.apply(extract_year, axis=1)
    day_loc_conc_avg1['month'] = day_loc_conc_avg1.apply(extract_month, axis=1)

    return day_loc_conc_avg1


my_data = so2_df2_2012_plus

day_loc_conc_avg1 = create_day_level(my_data)
day_level_location_map = location_wise_data(day_loc_conc_avg1)
location_day_anonaly_map = {}
location_month_map = {}
for location in day_level_location_map.keys():
    location_day_anonaly_map[location] = mark_outliers(
        day_level_location_map[location], 'mean')
    plt.figure(figsize=(20, 8))
    lp = sns.scatterplot(data=location_day_anonaly_map[location], x='Date', y='mean', hue='is_outlier', palette={
                         True: 'red', False: 'blue'})
    lp.set_title(location, size=16)
    plt.show()

    loc_month_df = aggregate_for_month(day_level_location_map[location])
    location_month_map[location] = mark_outliers(loc_month_df, 'mean', 3)
    cp = sns.catplot(data=location_month_map[location], kind='box', x='month', y='mean',
                     sharey=False, height=4, aspect=2)  # ,hue='is_outlier',palette={True: 'red', False: 'blue'})
    cp.fig.suptitle(location, size=16)
    plt.show()


new_df = pd.DataFrame(so2_df2_2012_plus)
new_df.to_excel("main dataset.xlsx", index=False)

day_loc_conc_avg1.pivot(
    columns='location', values='mean').plot.hist(bins=100, alpha=0.7)

plt.show()


day_loc_count = so2_df2_2012_plus.groupby(['Date', 'location']).size()


fig, ax = plt.subplots(figsize=(9, 7))
sns.violinplot(
    ax=ax, x=day_loc_conc_avg1["location"], y=day_loc_conc_avg1["mean"], palette='Set1')

sns.lineplot(data=day_loc_conc_avg1,
             x=day_loc_conc_avg1["Date"], y=day_loc_conc_avg1["mean"], hue='location')
plt.show()


annual_so2 = so2_df2_2012_plus.groupby(['year', 'location'])['Conc.'].agg([
    np.sum, np.mean, np.std, np.median])
annual_so2 = annual_so2.reset_index()

sns.lineplot(data=annual_so2,
             x=annual_so2['year'], y=annual_so2["mean"], hue='location')
plt.show()

sns.lineplot(data=annual_so2,
             x=annual_so2['year'], y=annual_so2["median"], hue='location')
plt.show()


monthly_so2 = so2_df2_2012_plus.groupby(['year', 'month', 'location'])[
    'Conc.'].agg([np.sum, np.mean, np.std, np.median])
monthly_so2 = monthly_so2.reset_index()
monthly_so2['month_date'] = monthly_so2.apply(create_date_for_month, axis=1)


fig, ax = plt.subplots(figsize=(20, 7))
sns.lineplot(data=monthly_so2,
             x=monthly_so2['month_date'], y=monthly_so2["mean"], hue='location')
plt.show()

df = monthly_so2
df['moving'] = df.groupby('location')['mean'].transform(
    lambda x: x.rolling(10, 1).mean())

fig, ax = plt.subplots(figsize=(20, 7))
sns.lineplot(data=monthly_so2,
             x=monthly_so2['month_date'], y=monthly_so2["moving"], hue='location')
plt.show()


ku_so2 = df[df.location == 'Kupwad']
rc_so2 = df[df.location == 'Rajawada chowk']

ub_so2 = df[df.location == 'Udyoga Bhawan']


so2_v3_na = so2_df[so2_df['V(m3)'].isna()]
so2_abs_na = so2_df[so2_df['Abs.'].isna()]


tmp_sp2 = so2_list[0]
cols = tmp_sp2.columns

my_data = so2_list[0]
my_data1 = fill_cells(my_data)


def get_mode_freq(my_data):

    date_col = my_data[my_data.columns[0]]
    date_col = date_col.dropna()
    gap_list = []
    for a in range(1, date_col.size):
        gap = date_col.index[a] - date_col.index[a-1]
        gap_list.append(gap)
    gap_series = pd.Series(gap_list)
    val_counts = gap_series.value_counts()
    sorted_val_counts = val_counts.sort_values(ascending=False)
    #periodicity = sorted_val_counts[0]
    #mode = gap_series.mode()
    print(sorted_val_counts)
    return sorted_val_counts.index[0]


freq = get_mode_freq(my_data)


tmp_sp2.iloc[:, 1:4]


ub_so2 = my_excel.iloc[2:, 0:5]
ub_so2.columns = ub_so2.iloc[0]
ub_so2.drop(index=2, axis=0, inplace=True)

ku_so2 = my_excel.iloc[:, 5:10]
rc_so2 = my_excel.iloc[:, 10:15]

ub_nox = my_excel.iloc[:, 15:21]
ku_nox = my_excel.iloc[:, 21:27]
rc_nox = my_excel.iloc[:, 27:33]


def get_row_data(row):
    row_map = {}
    for j in range(len(row)):
        value = row[j].value

        if not (row[j].value == None):
            print(j, value)
            row_map[j] = value


i = 0
for row in sheet.iter_rows():
    print("new_row", i)
    get_row_data(row)
    if(row[0].value == 'Date'):
        break
    i += 1


# for j in range(len(row)):
#    print(row[j].value)

# Iterate through merged cells
for merged_range in sheet.merged_cells.ranges:
    # Get the top-left cell of the merged range
    top_left_cell = merged_range.start_cell
    value = top_left_cell.value  # Value of the merged cell

    # Fill all cells in the merged range with the top-left value
    for row in sheet[merged_range.coord]:
        for cell in row:
            cell.value = value

# Save the updated workbook


path = "E:/Air/Air pollution/Annual/2024/jANUARY 2024"
air_file = path+"/Template_Air.xls"
dust_file = path+"/TemplateDust.xls"

air_xslx = air_file.split('.')[0] + ".xlsx"

# Convert .xls to .xlsx
pyexcel.save_book_as(file_name=air_file, dest_file_name=air_xslx)
workbook = load_workbook(air_xslx)

year_str = '2013'


workbook.save("updated_example.xlsx")
